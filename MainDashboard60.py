#!/usr/bin/env python3
"""
Generates an Excel dashboard with:
- Risk On/Off signals in column A for every metric
- USDT Dominance (CMC + local cache)
- Fear & Greed Index (Alternative.me) + streaks
- Golden Cross and EMA Cross for SPX, VIX, DXY, BTC
- BTC Aggregate Funding Rate (Coinalyze)
- BTC ETF inflows (SoSoValue)
- CBBI (ColinTalksCrypto Bull Run Index) + underlying components (Pi Cycle, RHODL, etc.)
- Weighted Summary in A2 (CBBI weighted highest)
"""

# ===========================================================
# 1. Imports, keys, configuration
# ===========================================================
import sys, time, os, re, subprocess, importlib
from datetime import datetime, timezone, timedelta
import json

CMC_API_KEY = "13314f52fac04e1aa455d91af3604b48"
CMC_BASE = "https://pro-api.coinmarketcap.com/v1"

COINALYZE_API_KEY = "2b6a93dc-98d7-4904-9994-dd7430d34909"

SOSO_API_KEY = "SOSO-814f6b10c77842ab8d5df65353ac1ce6"
SOSO_BASE = "https://openapi.sosovalue.com/api/v1"

CACHE_FILE = "usdt_dominance_cache.json"
CBBI_CACHE_FILE = "cbbi_cache.json"

DATE_STR = datetime.now(timezone.utc).strftime("%m%d%Y")
FILENAME = f"Fear_Greed_Dashboard_{DATE_STR}.xlsx"
SHEET_NAME = "Dashboard"

DESIRED_HEADER = ["Metric", "Value", "", "", "", "", "Legend", "Data Source", "Note"]
COL_WIDTHS = {"A": 15, "B": 42, "C": 42, "D": 20, "E": 20, "F": 20, "G": 20, "H": 85, "I": 20, "J": 85}
MAX_COL_WIDTH = 120

# ===========================================================
# 2. Ensure required packages
# ===========================================================
REQ = [
    "selenium",
    "webdriver-manager",
    "openpyxl",
    "requests",
    "pandas",
    "yfinance",
    "coinalyze",
    "beautifulsoup4",
]

def ensure(pkgs):
    miss = []
    for p in pkgs:
        if p == "beautifulsoup4":
            modname = "bs4"
        elif p.startswith("webdriver-manager"):
            modname = "webdriver_manager"
        else:
            modname = p

        base_mod = modname.split("==")[0]

        try:
            importlib.import_module(base_mod)
        except Exception:
            miss.append(p)

    if miss:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade"] + miss)
        time.sleep(1)

    for p in pkgs:
        if p == "beautifulsoup4":
            modname = "bs4"
        elif p.startswith("webdriver-manager"):
            modname = "webdriver_manager"
        else:
            modname = p

        base_mod = modname.split("==")[0]

        try:
            importlib.import_module(base_mod)
        except Exception as e:
            print("Missing package after install attempt:", p, e)
            return False

    return True

if not ensure(REQ):
    print("Please enable pip installs or preinstall packages:", REQ)
    sys.exit(1)

# ===========================================================
# 3. External libraries
# ===========================================================
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
import pandas as pd
import yfinance as yf
from coinalyze import CoinalyzeClient, HistoryEndpoint, Interval
from bs4 import BeautifulSoup

# ===========================================================
# 4. Styling and workbook helpers
# ===========================================================
GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL   = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
CLEAR_FILL = PatternFill()

def apply_column_styles(ws):
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(size=13, bold=True)
    thin_bottom = Border(bottom=Side(style="thin"))

    a1 = ws.cell(row=1, column=1)
    b1 = ws.cell(row=1, column=2)
    if a1.value:
        a1.font = bold_font
        a1.alignment = center
    if b1.value:
        b1.font = bold_font
        b1.alignment = center

    for r in range(2, ws.max_row + 1):
        for c in range(2, 8):
            ws.cell(row=r, column=c).alignment = center
            ws.cell(row=r, column=c).font = Font(bold=(c == 2), size=13)
        for c in (8, 9, 10):
            ws.cell(row=r, column=c).alignment = left_wrap
            ws.cell(row=r, column=c).font = Font(size=13)
        for c in range(2, 11):
            ws.cell(row=r, column=c).border = Border()
        metric_name = (ws.cell(row=r, column=2).value or "").strip().lower()
        wants_bottom = metric_name in ["usdt dominance – historical", "fear & greed index"] or metric_name.startswith("ema cross")
        if wants_bottom:
            for c in range(2, 11):
                ws.cell(row=r, column=c).border = thin_bottom

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = min(width, MAX_COL_WIDTH)

def ensure_workbook_header(filename, sheet_name):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)
        wb.close()
        return
    wb = load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    wb.save(filename)
    wb.close()

def create_new_workbook(filename, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    eastern = timezone(timedelta(hours=-5))
    now_et = datetime.now(eastern)
    today_str = now_et.strftime("%m/%d/%Y")
    raw_time = now_et.strftime("%I:%M%p").lower()
    time_str = raw_time.lstrip("0")

    ws.cell(row=1, column=1, value=today_str).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=2, value=time_str).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filename)
    wb.close()

# ===========================================================
# 5. Price series and crosses
# ===========================================================
def fetch_price_series_yf(symbol):
    ticker = yf.Ticker(symbol)
    hist = ticker.history(period="max", auto_adjust=False)
    if hist.empty:
        raise RuntimeError(f"No data for {symbol}")
    df = hist.reset_index()[["Date", "Close"]].rename(columns={"Date": "date", "Close": "close"})
    df["date"] = pd.to_datetime(df["date"]).dt.date
    return df.dropna(subset=["close"]).drop_duplicates("date").sort_values("date").reset_index(drop=True)

def compute_golden_cross_status_from_price_df(df_price, fast_len=50, slow_len=200):
    df = df_price.copy().sort_values("date").reset_index(drop=True)
    df["fast"] = df["close"].rolling(window=fast_len, min_periods=fast_len).mean()
    df["slow"] = df["close"].rolling(window=slow_len, min_periods=slow_len).mean()
    df = df.dropna(subset=["fast", "slow"]).reset_index(drop=True)
    latest = df.iloc[-1]
    fast_val = float(latest["fast"])
    slow_val = float(latest["slow"])
    gap_pct = (fast_val / slow_val - 1.0) * 100.0 if slow_val else float("nan")
    df["diff"] = df["fast"] - df["slow"]

    def sign_of(x):
        return 1 if x > 0 else (-1 if x < 0 else 0)

    df["sign"] = df["diff"].apply(sign_of)
    df["sign_prev"] = df["sign"].shift(1)
    crosses = df[(df["sign"].notna()) & (df["sign_prev"].notna()) & (df["sign"] != df["sign_prev"])]
    cycles = len(crosses)
    last_cross_date, last_cross_type = None, None
    if not crosses.empty:
        row = crosses.iloc[-1]
        last_cross_date = pd.to_datetime(row["date"]).date().isoformat()
        last_cross_type = "Golden" if row["sign"] > row["sign_prev"] else "Death"
    state = "Golden" if fast_val > slow_val else ("Death" if fast_val < slow_val else "Neutral")
    days_since = None
    if last_cross_date:
        dlast = pd.to_datetime(last_cross_date).date()
        dtoday = pd.to_datetime("now", utc=True).date()
        days_since = (dtoday - dlast).days
    recent_cross = None
    if len(df) >= 2:
        s_prev = df["sign"].iat[-2]
        s_now = df["sign"].iat[-1]
        if s_prev != s_now:
            recent_cross = "Golden" if s_now > s_prev else "Death"
    return {
        "state": state,
        "fast": fast_val,
        "slow": slow_val,
        "gap_pct": gap_pct,
        "last_cross_date": last_cross_date,
        "last_cross_type": last_cross_type,
        "cycles": cycles,
        "days_since_last_cross": days_since,
        "recent_cross": recent_cross,
    }

def compute_ema_cross_status_from_price_df(df_price, fast_len=21, slow_len=55):
    df = df_price.copy().sort_values("date").reset_index(drop=True)
    df["fast"] = df["close"].ewm(span=fast_len, adjust=False).mean()
    df["slow"] = df["close"].ewm(span=slow_len, adjust=False).mean()
    df = df.dropna(subset=["fast", "slow"]).reset_index(drop=True)
    latest = df.iloc[-1]
    fast_val = float(latest["fast"])
    slow_val = float(latest["slow"])
    gap_pct = (fast_val / slow_val - 1.0) * 100.0 if slow_val else float("nan")
    df["diff"] = df["fast"] - df["slow"]

    def sign_of(x):
        return 1 if x > 0 else (-1 if x < 0 else 0)

    df["sign"] = df["diff"].apply(sign_of)
    df["sign_prev"] = df["sign"].shift(1)
    crosses = df[(df["sign"].notna()) & (df["sign_prev"].notna()) & (df["sign"] != df["sign_prev"])]
    cycles = len(crosses)
    last_cross_date, last_cross_type = None, None
    if not crosses.empty:
        row = crosses.iloc[-1]
        last_cross_date = pd.to_datetime(row["date"]).date().isoformat()
        last_cross_type = "Golden" if row["sign"] > row["sign_prev"] else "Death"
    state = "Golden" if fast_val > slow_val else ("Death" if fast_val < slow_val else "Neutral")
    days_since = None
    if last_cross_date:
        dlast = pd.to_datetime(last_cross_date).date()
        dtoday = pd.to_datetime("now", utc=True).date()
        days_since = (dtoday - dlast).days
    recent_cross = None
    if len(df) >= 2:
        s_prev = df["sign"].iat[-2]
        s_now = df["sign"].iat[-1]
        if s_prev != s_now:
            recent_cross = "Golden" if s_now > s_prev else "Death"
    return {
        "state": state,
        "fast": fast_val,
        "slow": slow_val,
        "gap_pct": gap_pct,
        "last_cross_date": last_cross_date,
        "last_cross_type": last_cross_type,
        "cycles": cycles,
        "days_since_last_cross": days_since,
        "recent_cross": recent_cross,
    }

def apply_cross_risk_logic(ws, row, metric_name, cross_state, golden_means_on=True):
    label, fill = None, None
    if cross_state == "Golden":
        label, fill = ("⬆️ Risk On", GREEN_FILL) if golden_means_on else ("⬇️ Risk Off", RED_FILL)
    elif cross_state == "Death":
        label, fill = ("⬇️ Risk Off", RED_FILL) if golden_means_on else ("⬆️ Risk On", GREEN_FILL)
    else:
        label = "Neutral"
    cell = ws.cell(row=row, column=1, value=label)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ===========================================================
# 6. Cell writers and conditional coloring
# ===========================================================
def color_value_cell(ws, row, col, value, metric_name=None):
    cell = ws.cell(row=row, column=col)
    sval = str(value) if value is not None else ""
    s_low = sval.lower()
    if "golden" in s_low:
        cell.fill = GREEN_FILL
        return
    if "death" in s_low:
        cell.fill = RED_FILL
        return
    if metric_name and metric_name.strip().lower() == "fear & greed index":
        try:
            num = float(value)
            if num < 30:
                cell.fill = RED_FILL
                return
            if num > 70:
                cell.fill = GREEN_FILL
                return
        except Exception:
            pass
    cell.fill = PatternFill()

def _format_date_string_mmddyyyy(maybe_date):
    if maybe_date is None:
        return None
    try:
        if isinstance(maybe_date, datetime):
            return maybe_date.strftime("%m/%d/%Y")
        if isinstance(maybe_date, str):
            parsed = pd.to_datetime(maybe_date, errors="coerce")
            if not pd.isna(parsed):
                return parsed.strftime("%m/%d/%Y")
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", maybe_date.strip())
            if m:
                mon = int(m.group(1))
                day = int(m.group(2))
                yr = int(m.group(3))
                return f"{mon:02d}/{day:02d}/{yr}"
    except Exception:
        return None
    return None

def upsert_golden_cross_row(filename, sheet_name, row_number, metric_name, status_dict, source="yfinance daily; 50/200 SMA"):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    while ws.max_row < row_number:
        ws.append([None] * len(DESIRED_HEADER))
    ws.cell(row=row_number, column=2, value=metric_name)
    state_val = status_dict.get("state")
    ws.cell(row=row_number, column=3, value=state_val)
    lcd_fmt = _format_date_string_mmddyyyy(status_dict.get("last_cross_date"))
    ws.cell(row=row_number, column=4, value=lcd_fmt)
    ws.cell(row=row_number, column=5, value=status_dict.get("cycles"))
    ws.cell(row=row_number, column=6, value=status_dict.get("days_since_last_cross"))
    cell_f = ws.cell(row=row_number, column=7, value=round(status_dict.get("gap_pct", 0.0), 2))
    cell_f.number_format = "0.00"
    ws.cell(row=row_number, column=8, value="Value=state; C=last_cross_date; D=cycles; E=days_since_last_cross; F=gap_pct; G=legend; H=source; I=note")
    ws.cell(row=row_number, column=9, value=source)
    note = f"fast={status_dict.get('fast'):.2f}; slow={status_dict.get('slow'):.2f}; last_cross_type={status_dict.get('last_cross_type')}; recent_cross={status_dict.get('recent_cross')}"
    ws.cell(row=row_number, column=10, value=note)
    color_value_cell(ws, row_number, 3, state_val, metric_name=metric_name)
    wb.save(filename)
    wb.close()

def upsert_ema_cross_row(filename, sheet_name, row_number, metric_name, status_dict, source="yfinance daily; 21/55 EMA"):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    while ws.max_row < row_number:
        ws.append([None] * len(DESIRED_HEADER))
    ws.cell(row=row_number, column=2, value=metric_name)
    state_val = status_dict.get("state")
    ws.cell(row=row_number, column=3, value=state_val)
    lcd_fmt = _format_date_string_mmddyyyy(status_dict.get("last_cross_date"))
    ws.cell(row=row_number, column=4, value=lcd_fmt)
    ws.cell(row=row_number, column=5, value=status_dict.get("cycles"))
    ws.cell(row=row_number, column=6, value=status_dict.get("days_since_last_cross"))
    cell_f = ws.cell(row=row_number, column=7, value=round(status_dict.get("gap_pct", 0.0), 2))
    cell_f.number_format = "0.00"
    ws.cell(row=row_number, column=8, value="Value=state; C=last_cross_date; D=cycles; E=days_since_last_cross; F=gap_pct; G=legend; H=source; I=note")
    ws.cell(row=row_number, column=9, value=source)
    note = f"fast={status_dict.get('fast'):.2f}; slow={status_dict.get('slow'):.2f}; last_cross_type={status_dict.get('last_cross_type')}; recent_cross={status_dict.get('recent_cross')}"
    ws.cell(row=row_number, column=10, value=note)
    color_value_cell(ws, row_number, 3, state_val, metric_name=metric_name)
    wb.save(filename)
    wb.close()

# ===========================================================
# 7. Risk labels for historical rows
# ===========================================================
def apply_risk_label(ws, row, col, now_val, yesterday_val, month_val, higher_means_off=True):
    label, fill = "Neutral", None
    if now_val is not None and month_val is not None:
        if higher_means_off:
            if now_val > month_val:
                label, fill = "⬇️ Risk Off", RED_FILL
            elif yesterday_val is not None and now_val < month_val and yesterday_val < month_val:
                label, fill = "⬆️ Risk On", GREEN_FILL
        else:
            if now_val < month_val:
                label, fill = "⬇️ Risk Off", RED_FILL
            elif yesterday_val is not None and now_val > month_val and yesterday_val > month_val:
                label, fill = "⬆️ Risk On", GREEN_FILL
    cell = ws.cell(row=row, column=col, value=label)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ===========================================================
# 8. Data sources: USDT dominance (CMC) + cache
# ===========================================================
def fetch_cmc_global_and_usdt():
    headers = {"X-CMC_PRO_API_KEY": CMC_API_KEY}
    resp_g = requests.get(f"{CMC_BASE}/global-metrics/quotes/latest", headers=headers, timeout=15)
    resp_g.raise_for_status()
    total_cap = resp_g.json()["data"]["quote"]["USD"]["total_market_cap"]
    resp_u = requests.get(f"{CMC_BASE}/cryptocurrency/quotes/latest?id=825", headers=headers, timeout=15)
    resp_u.raise_for_status()
    usdt_cap = resp_u.json()["data"]["825"]["quote"]["USD"]["market_cap"]
    return total_cap, usdt_cap

def update_cache(today_value):
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cache = {}
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r") as f:
            cache = json.load(f)
    cache[today_str] = today_value
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f)

def lookup_cache(days_back):
    target = datetime.now(timezone.utc).date() - timedelta(days=days_back)
    if not os.path.exists(CACHE_FILE):
        return None
    with open(CACHE_FILE, "r") as f:
        cache = json.load(f)
    entries = [(datetime.fromisoformat(d).date(), cache[d]) for d in cache]
    entries.sort(key=lambda x: x[0])
    candidates = [val for (dt, val) in entries if dt <= target]
    return candidates[-1] if candidates else None

# ===========================================================
# 9. Data sources: Fear & Greed (Alternative.me)
# ===========================================================
def fetch_fng_series_altme(limit=730):
    url = f"https://api.alternative.me/fng/?limit={limit}&format=json"
    r = requests.get(url, timeout=15)
    r.raise_for_status()
    data = r.json()
    return [int(item["value"]) for item in data.get("data", [])]

# ===========================================================
# 10. Data sources: Coinalyze funding rates
# ===========================================================
def fetch_btc_funding_rate():
    client = CoinalyzeClient(api_key=COINALYZE_API_KEY)
    resp = client.get_history(
        endpoint=HistoryEndpoint.FUNDING_RATE,
        symbols="BTCUSDT_PERP.A",
        interval=Interval.D1,
    )
    if not resp or "history" not in resp[0]:
        return {"now": None, "yesterday": None, "week": None, "month": None, "year_avg": None}
    history = resp[0]["history"]
    if not history:
        return {"now": None, "yesterday": None, "week": None, "month": None, "year_avg": None}
    latest = history[-1]["c"]
    yesterday = history[-2]["c"] if len(history) > 1 else None
    week = history[-7]["c"] if len(history) > 7 else None
    month = history[-30]["c"] if len(history) > 30 else None
    year_values = [h["c"] for h in history[-365:] if h.get("c") is not None]
    year_avg = round(sum(year_values) / len(year_values), 6) if year_values else None
    return {
        "now": round(latest, 6) if latest is not None else None,
        "yesterday": round(yesterday, 6) if yesterday is not None else None,
        "week": round(week, 6) if week is not None else None,
        "month": round(month, 6) if month is not None else None,
        "year_avg": year_avg,
    }

# ===========================================================
# 10b. Data sources: CBBI + components (scrape + cache)
# ===========================================================
CBBI_COMPONENT_LABELS = [
    "Pi Cycle Top Indicator",
    "RUPL / NUPL Chart",
    "RHODL Ratio",
    "Puell Multiple",
    "2 Year Moving Average",
    "Bitcoin Trolololo Trend Line",
    "MVRV Z-Score",
    "Reserve Risk",
    "Woobull Top Cap vs CVDD",
]

def _load_cbbi_cache():
    if not os.path.exists(CBBI_CACHE_FILE):
        return {}
    try:
        with open(CBBI_CACHE_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_cbbi_cache(cache):
    with open(CBBI_CACHE_FILE, "w") as f:
        json.dump(cache, f)

def _extract_metric_from_label_text(text):
    m = re.search(r"(\d+(\.\d+)?)\s*%", text)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None

def _scrape_cbbi_page():
    url = "https://colintalkscrypto.com/cbbi/"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://google.com",
        "Connection": "keep-alive",
    }
    r = requests.get(url, headers=headers, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    metrics = {}
    for label in CBBI_COMPONENT_LABELS:
        el = soup.find(string=re.compile(re.escape(label), re.IGNORECASE))
        if not el:
            metrics[label] = None
            continue
        text = el.parent.get_text(" ", strip=True)
        val = _extract_metric_from_label_text(text)
        metrics[label] = val

    valid_vals = [v for v in metrics.values() if isinstance(v, (int, float))]
    cbbi_now = round(sum(valid_vals) / len(valid_vals), 2) if valid_vals else None
    return cbbi_now, metrics

def update_cbbi_cache_for_today():
    cache = _load_cbbi_cache()
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cbbi_now, metrics = _scrape_cbbi_page()
    cache[today_str] = {"cbbi": cbbi_now, "metrics": metrics}
    _save_cbbi_cache(cache)
    return today_str, cbbi_now, metrics, cache

def _series_from_cbbi_cache(cache, key_func):
    entries = []
    for d, entry in cache.items():
        try:
            dt = datetime.fromisoformat(d).date()
        except Exception:
            continue
        val = key_func(d, entry)
        if val is not None:
            entries.append((dt, float(val)))
    if not entries:
        return {"now": None, "yesterday": None, "week": None, "month": None, "year_avg": None}

    entries.sort(key=lambda x: x[0])
    last_date = entries[-1][0]

    def closest_on_or_before(days):
        target = last_date - timedelta(days=days)
        candidates = [v for dt, v in entries if dt <= target]
        return candidates[-1] if candidates else None

    now = entries[-1][1]
    yesterday = closest_on_or_before(1)
    week = closest_on_or_before(7)
    month = closest_on_or_before(30)
    year_cut = last_date - timedelta(days=365)
    year_vals = [v for dt, v in entries if dt >= year_cut]
    year_avg = round(sum(year_vals) / len(year_vals), 2) if year_vals else None

    return {"now": now, "yesterday": yesterday, "week": week, "month": month, "year_avg": year_avg}

def compute_cbbi_series_from_cache(cache):
    return _series_from_cbbi_cache(cache, key_func=lambda d, entry: entry.get("cbbi"))

def compute_cbbi_component_series_from_cache(cache, component_label):
    return _series_from_cbbi_cache(
        cache,
        key_func=lambda d, entry: (entry.get("metrics") or {}).get(component_label),
    )

def upsert_cbbi_row(filename, sheet_name, row_number, cbbi_data, weight=3):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    while ws.max_row < row_number:
        ws.append([None] * len(DESIRED_HEADER))

    ws.cell(row=row_number, column=2, value="CBBI – Historical")

    now_val = cbbi_data.get("now")
    y_val = cbbi_data.get("yesterday")
    w_val = cbbi_data.get("week")
    m_val = cbbi_data.get("month")
    avg_val = cbbi_data.get("year_avg")

    ws.cell(row=row_number, column=3, value=now_val if now_val is not None else "Unavailable")
    ws.cell(row=row_number, column=4, value=y_val if y_val is not None else "Unavailable")
    ws.cell(row=row_number, column=5, value=w_val if w_val is not None else "Unavailable")
    ws.cell(row=row_number, column=6, value=m_val if m_val is not None else "Unavailable")
    ws.cell(row=row_number, column=7, value=avg_val if avg_val is not None else "Unavailable")

    ws.cell(row=row_number, column=8, value="B=now; C=yesterday; D=week; E=month; F=year avg")
    ws.cell(row=row_number, column=9, value="Scraped from colintalkscrypto.com/cbbi + local cache")
    ws.cell(row=row_number, column=10, value=f"Composite CBBI (approx from components); Weight in summary = {weight}")

    for col in range(3, 8):
        cell = ws.cell(row=row_number, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"

    apply_risk_label(
        ws,
        row=row_number,
        col=1,
        now_val=now_val,
        yesterday_val=y_val,
        month_val=m_val,
        higher_means_off=True,
    )

    wb.save(filename)
    wb.close()

def upsert_cbbi_component_row(filename, sheet_name, row_number, label, series_data, weight=1):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    while ws.max_row < row_number:
        ws.append([None] * len(DESIRED_HEADER))

    metric_name = f"CBBI {label} – Historical"
    ws.cell(row=row_number, column=2, value=metric_name)

    now_val = series_data.get("now")
    y_val = series_data.get("yesterday")
    w_val = series_data.get("week")
    m_val = series_data.get("month")
    avg_val = series_data.get("year_avg")

    ws.cell(row=row_number, column=3, value=now_val if now_val is not None else "Unavailable")
    ws.cell(row=row_number, column=4, value=y_val if y_val is not None else "Unavailable")
    ws.cell(row=row_number, column=5, value=w_val if w_val is not None else "Unavailable")
    ws.cell(row=row_number, column=6, value=m_val if m_val is not None else "Unavailable")
    ws.cell(row=row_number, column=7, value=avg_val if avg_val is not None else "Unavailable")

    ws.cell(row=row_number, column=8, value="B=now; C=yesterday; D=week; E=month; F=year avg")
    ws.cell(row=row_number, column=9, value="Scraped from colintalkscrypto.com/cbbi + local cache")
    ws.cell(row=row_number, column=10, value=f"CBBI component: {label}; Weight in summary = {weight}")

    for col in range(3, 8):
        cell = ws.cell(row=row_number, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"

    apply_risk_label(
        ws,
        row=row_number,
        col=1,
        now_val=now_val,
        yesterday_val=y_val,
        month_val=m_val,
        higher_means_off=True,
    )

    wb.save(filename)
    wb.close()

# ===========================================================
# 12. Data sources: ETF inflows (SoSoValue) + row writer
# ===========================================================
def fetch_etf_inflows():
    headers = {"Authorization": f"Bearer {SOSO_API_KEY}"}
    url = f"{SOSO_BASE}/etf/btc-spot/inflows"
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    payload = resp.json()
    data = payload.get("data", [])
    if not isinstance(data, list) or not data:
        return {"now": None, "yesterday": None, "week": None, "month": None}
    today_inflow = sum(float(item.get("netInflow", 0) or 0) for item in data)
    return {"now": today_inflow, "yesterday": None, "week": None, "month": None}

def upsert_etf_inflows_row(filename, sheet_name, row_number, etf_data):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    while ws.max_row < row_number:
        ws.append([None] * len(DESIRED_HEADER))
    ws.cell(row=row_number, column=2, value="BTC ETF Inflows – Historical")
    now_val = etf_data.get("now")
    y_val = etf_data.get("yesterday")
    w_val = etf_data.get("week")
    m_val = etf_data.get("month")
    ws.cell(row=row_number, column=3, value=now_val if now_val is not None else "Unavailable")
    ws.cell(row=row_number, column=4, value=y_val if y_val is not None else "Unavailable")
    ws.cell(row=row_number, column=5, value=w_val if w_val is not None else "Unavailable")
    ws.cell(row=row_number, column=6, value=m_val if m_val is not None else "Unavailable")
    ws.cell(row=row_number, column=8, value="B=now; C=yesterday; D=last week; E=last month")
    ws.cell(row=row_number, column=9, value="SoSoValue API")
    ws.cell(row=row_number, column=10, value="Daily net inflows for US spot BTC ETFs")
    for col in range(3, 7):
        cell = ws.cell(row=row_number, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"
    apply_risk_label(ws, row=row_number, col=1, now_val=now_val, yesterday_val=y_val, month_val=m_val, higher_means_off=False)
    wb.save(filename)
    wb.close()

# ===========================================================
# 13. Precompute statuses
# ===========================================================
def precompute_statuses():
    df_spx = fetch_price_series_yf("SPY")
    df_vix = fetch_price_series_yf("^VIX")
    df_dxy = fetch_price_series_yf("DX-Y.NYB")
    df_btc = fetch_price_series_yf("BTC-USD")
    return {
        "df_spx": df_spx,
        "df_vix": df_vix,
        "df_dxy": df_dxy,
        "df_btc": df_btc,
        "status_spx_gc": compute_golden_cross_status_from_price_df(df_spx),
        "status_vix_gc": compute_golden_cross_status_from_price_df(df_vix),
        "status_dxy_gc": compute_golden_cross_status_from_price_df(df_dxy),
        "status_btc_gc": compute_golden_cross_status_from_price_df(df_btc),
        "status_spx": compute_ema_cross_status_from_price_df(df_spx),
        "status_vix": compute_ema_cross_status_from_price_df(df_vix),
        "status_dxy": compute_ema_cross_status_from_price_df(df_dxy),
        "status_btc": compute_ema_cross_status_from_price_df(df_btc),
    }

# ===========================================================
# 14. Risk summary (A2) — weighted and color-coded
# ===========================================================
def update_risk_summary(ws):
    risk_on_score, risk_off_score = 0, 0
    weights = {
        "CBBI – Historical": 3,
        "Golden Cross BTC-USD 50/200 daily": 2,
        "EMA Cross BTC-USD 21/55 daily": 2,
        "Golden Cross S&P500 50/200 daily": 2,
        "EMA Cross S&P500 21/55 daily": 2,
        "Golden Cross VIX 50/200 daily": 1,
        "EMA Cross VIX 21/55 daily": 1,
        "Golden Cross DXY 50/200 daily": 1,
        "EMA Cross DXY 21/55 daily": 1,
        "BTC Aggregate Funding Rate – Historical": 1,
        "Fear & Greed – Historical": 1,
        "USDT Dominance – Historical": 1,
        "BTC ETF Inflows – Historical": 1,
    }

    for label in CBBI_COMPONENT_LABELS:
        metric_name = f"CBBI {label} – Historical"
        weights[metric_name] = 1

    for r in range(3, ws.max_row + 1):
        metric = (ws.cell(row=r, column=2).value or "").strip()
        val = (ws.cell(row=r, column=1).value or "").strip().lower()
        w = weights.get(metric, 1)
        if "risk on" in val:
            risk_on_score += w
        elif "risk off" in val:
            risk_off_score += w

    diff = risk_on_score - risk_off_score
    summary = (
        f"{diff} Weighted Risk On" if diff > 0
        else f"{abs(diff)} Weighted Risk Off" if diff < 0
        else "Neutral"
    )

    cell = ws.cell(row=2, column=1, value=summary)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = RED_FILL if "risk off" in summary.lower() else GREEN_FILL
    thick = Side(style="thick")
    cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)

    ws.cell(
        row=2,
        column=8,
        value="Summary: weighted Risk On minus Risk Off; CBBI and components weighted as noted in their rows",
    )
# ===========================================================
# 15. Main execution
# ===========================================================
if __name__ == "__main__":
    # Ensure workbook exists before any loads
    create_new_workbook(FILENAME, SHEET_NAME)
    ensure_workbook_header(FILENAME, SHEET_NAME)

    # Fear & Greed history
    history_values = []
    fear_greed_value = None
    try:
        history_values = fetch_fng_series_altme(limit=730)
        if history_values:
            fear_greed_value = history_values[0]
    except Exception as e:
        print("Error fetching Fear & Greed Index:", e)

    wb = load_workbook(FILENAME)
    ws = wb[SHEET_NAME]

    # Motivational line
    ws.cell(row=2, column=2, value="First you Survive and then you Thrive")
    ws.cell(row=2, column=3, value="You're a deep value guy, remember that. So buy at the beginning of RISK ON")
    ws.cell(row=2, column=4, value="When are the SELLERS EXHAUSTED? And when are the BUYERS BUYING?")
    ws.cell(row=2, column=5, value="RANGING Market OR TRENDING Market?")
    for c in (2, 3):
        ws.cell(row=2, column=c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=2, column=c).font = Font(size=12, bold=(c == 2))
    wb.save(FILENAME)
    wb.close()

    # USDT Dominance – Historical (row 3)
    try:
        total_cap, usdt_cap = fetch_cmc_global_and_usdt()
        usdt_d_now = round((usdt_cap / total_cap) * 100, 2)
        update_cache(usdt_d_now)
    except Exception as e:
        print("Error fetching USDT dominance from CMC:", e)
        usdt_d_now = None

    usdt_d = {
        "now": usdt_d_now,
        "yesterday": lookup_cache(1),
        "week": lookup_cache(7),
        "month": lookup_cache(30),
    }

    wb = load_workbook(FILENAME)
    ws = wb[SHEET_NAME]
    ws.cell(row=3, column=2, value="USDT Dominance – Historical")
    ws.cell(row=3, column=3, value=usdt_d.get("now"))
    ws.cell(row=3, column=4, value=usdt_d.get("yesterday"))
    ws.cell(row=3, column=5, value=usdt_d.get("week"))
    ws.cell(row=3, column=6, value=usdt_d.get("month"))
    ws.cell(row=3, column=8, value="B=now; C=yesterday; D=last week; E=last month")
    ws.cell(row=3, column=9, value="CoinMarketCap Global Metrics API + Local Cache")
    ws.cell(row=3, column=10, value="Dominance = USDT cap / total crypto market cap; history cached locally")
    for col in range(3, 7):
        ws.cell(row=3, column=col).number_format = "0.00"
    color_value_cell(ws, 3, 3, usdt_d.get("now"), metric_name="USDT Dominance")
    apply_risk_label(
        ws,
        row=3,
        col=1,
        now_val=usdt_d.get("now"),
        yesterday_val=usdt_d.get("yesterday"),
        month_val=usdt_d.get("month"),
        higher_means_off=True,
    )
    wb.save(FILENAME)
    wb.close()

    # Fear & Greed – Historical (row 4)
    fg_now = history_values[0] if len(history_values) > 0 else None
    fg_yesterday = history_values[1] if len(history_values) > 1 else None
    fg_week = history_values[7] if len(history_values) > 7 else None
    fg_month = history_values[30] if len(history_values) > 30 else None

    wb = load_workbook(FILENAME)
    ws = wb[SHEET_NAME]
    ws.cell(row=4, column=2, value="Fear & Greed – Historical")
    ws.cell(row=4, column=3, value=fg_now)
    ws.cell(row=4, column=4, value=fg_yesterday)
    ws.cell(row=4, column=5, value=fg_week)
    ws.cell(row=4, column=6, value=fg_month)
    ws.cell(row=4, column=8, value="B=now; C=yesterday; D=last week; E=last month")
    ws.cell(row=4, column=9, value="Alternative.me API")
    ws.cell(row=4, column=10, value="Values from historical sentiment series; newest to oldest")
    color_value_cell(ws, 4, 3, fg_now, metric_name="Fear & Greed Index")
    apply_risk_label(
        ws,
        row=4,
        col=1,
        now_val=fg_now,
        yesterday_val=fg_yesterday,
        month_val=fg_month,
        higher_means_off=False,
    )
    wb.save(FILENAME)
    wb.close()

    # Fear & Greed streaks (row 5)
    longest_fear_streak = 0
    longest_greed_streak = 0
    current_fear_streak = 0
    current_greed_streak = 0
    pct_fear = None
    pct_greed = None

    if history_values:
        series = list(reversed(history_values))
        cur_fear, cur_greed = 0, 0
        fear_days, greed_days = 0, 0

        for v in series:
            if v < 31:
                cur_fear += 1
                longest_fear_streak = max(longest_fear_streak, cur_fear)
                fear_days += 1
                cur_greed = 0
            elif v > 69:
                cur_greed += 1
                longest_greed_streak = max(longest_greed_streak, cur_greed)
                greed_days += 1
                cur_fear = 0
            else:
                cur_fear = 0
                cur_greed = 0

        if fear_greed_value is not None and fear_greed_value < 31:
            current_fear_streak = cur_fear
        if fear_greed_value is not None and fear_greed_value > 69:
            current_greed_streak = cur_greed

        total_days = len(series)
        pct_fear = (fear_days / total_days * 100) if total_days else None
        pct_greed = (greed_days / total_days * 100) if total_days else None

    wb = load_workbook(FILENAME)
    ws = wb[SHEET_NAME]
    ws.cell(row=5, column=2, value="Fear & Greed Index")
    ws.cell(row=5, column=3, value=fear_greed_value)
    ws.cell(row=5, column=4, value=longest_fear_streak)
    ws.cell(row=5, column=5, value=longest_greed_streak)
    ws.cell(row=5, column=6, value=current_fear_streak)
    ws.cell(row=5, column=7, value=current_greed_streak)
    ws.cell(row=5, column=8, value=pct_fear)
    ws.cell(row=5, column=9, value=pct_greed)
    ws.cell(
        row=5,
        column=10,
        value="Source: Alternative.me API; B=current; C=longest<31; D=longest>69; E=current<31; F=current>69; G=%fear; H=%greed",
    )
    color_value_cell(ws, 5, 3, fear_greed_value, metric_name="Fear & Greed Index")
    apply_column_styles(ws)
    wb.save(FILENAME)
    wb.close()

    # Golden/EMA crosses (rows 6–13)
    statuses = precompute_statuses()
    upsert_golden_cross_row(FILENAME, SHEET_NAME, 6,  "Golden Cross S&P500 50/200 daily", statuses["status_spx_gc"])
    upsert_ema_cross_row   (FILENAME, SHEET_NAME, 7,  "EMA Cross S&P500 21/55 daily",    statuses["status_spx"])
    upsert_golden_cross_row(FILENAME, SHEET_NAME, 8,  "Golden Cross VIX 50/200 daily",   statuses["status_vix_gc"])
    upsert_ema_cross_row   (FILENAME, SHEET_NAME, 9,  "EMA Cross VIX 21/55 daily",       statuses["status_vix"])
    upsert_golden_cross_row(FILENAME, SHEET_NAME, 10, "Golden Cross DXY 50/200 daily",   statuses["status_dxy_gc"])
    upsert_ema_cross_row   (FILENAME, SHEET_NAME, 11, "EMA Cross DXY 21/55 daily",       statuses["status_dxy"])
    upsert_golden_cross_row(FILENAME, SHEET_NAME, 12, "Golden Cross BTC-USD 50/200 daily", statuses["status_btc_gc"])
    upsert_ema_cross_row   (FILENAME, SHEET_NAME, 13, "EMA Cross BTC-USD 21/55 daily",     statuses["status_btc"])

    wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
    apply_cross_risk_logic(ws, 6,  "Golden Cross S&P500", statuses["status_spx_gc"]["state"], golden_means_on=True)
    apply_cross_risk_logic(ws, 7,  "EMA Cross S&P500",    statuses["status_spx"]["state"],    golden_means_on=True)
    apply_cross_risk_logic(ws, 8,  "Golden Cross VIX",    statuses["status_vix_gc"]["state"], golden_means_on=False)
    apply_cross_risk_logic(ws, 9,  "EMA Cross VIX",       statuses["status_vix"]["state"],    golden_means_on=False)
    apply_cross_risk_logic(ws, 10, "Golden Cross DXY",    statuses["status_dxy_gc"]["state"], golden_means_on=False)
    apply_cross_risk_logic(ws, 11, "EMA Cross DXY",       statuses["status_dxy"]["state"],    golden_means_on=False)
    apply_cross_risk_logic(ws, 12, "Golden Cross BTC",    statuses["status_btc_gc"]["state"], golden_means_on=True)
    apply_cross_risk_logic(ws, 13, "EMA Cross BTC",       statuses["status_btc"]["state"],    golden_means_on=True)
    apply_column_styles(ws)
    wb.save(FILENAME); wb.close()

    # BTC Aggregate Funding Rate (row 14)
    try:
        btc_fr = fetch_btc_funding_rate()
    except Exception as e:
        print("Error fetching BTC funding rate:", e)
        btc_fr = {"now": None, "yesterday": None, "week": None, "month": None, "year_avg": None}
    wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
    ws.cell(row=14, column=2, value="BTC Aggregate Funding Rate – Historical")
    ws.cell(row=14, column=3, value=btc_fr.get("now"))
    ws.cell(row=14, column=4, value=btc_fr.get("yesterday"))
    ws.cell(row=14, column=5, value=btc_fr.get("week"))
    ws.cell(row=14, column=6, value=btc_fr.get("month"))
    ws.cell(row=14, column=7, value=btc_fr.get("year_avg"))
    ws.cell(row=14, column=8, value="C=now (red if ≤ month, else green); D=yesterday; E=last week; F=last month; G=year average")
    ws.cell(row=14, column=9, value="Coinalyze API (official client)")
    ws.cell(row=14, column=10, value="Daily funding rate history; G=average over last 365 days; Weight in summary = 1")
    apply_risk_label(ws, row=14, col=1, now_val=btc_fr.get("now"), yesterday_val=btc_fr.get("yesterday"), month_val=btc_fr.get("month"), higher_means_off=False)
    c_val = btc_fr.get("now"); f_val = btc_fr.get("month")
    c_cell = ws.cell(row=14, column=3)
    if c_val is not None and f_val is not None:
        c_cell.fill = RED_FILL if c_val <= f_val else GREEN_FILL
    c_cell.font = Font(size=13); c_cell.alignment = Alignment(horizontal="center", vertical="center")
    color_value_cell(ws, 14, 7, btc_fr.get("year_avg"), metric_name="BTC Funding Rate")
    for col in range(3, 8): ws.cell(row=14, column=col).number_format = "0.000000"
    apply_column_styles(ws)
    wb.save(FILENAME); wb.close()

    # ETF Inflows (row 16)
    try:
        etf_data = fetch_etf_inflows()
        upsert_etf_inflows_row(FILENAME, SHEET_NAME, 16, etf_data)
        wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
        c_cell = ws.cell(row=16, column=3); f_cell = ws.cell(row=16, column=6)
        c_val = c_cell.value; f_val = f_cell.value
        for col in range(3, 7):
            cell = ws.cell(row=16, column=col)
            if isinstance(cell.value, (int, float)): cell.number_format = "0.00"
        if isinstance(c_val, (int, float)) and isinstance(f_val, (int, float)):
            c_cell.fill = RED_FILL if c_val <= f_val else GREEN_FILL
        c_cell.font = Font(size=13); c_cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_column_styles(ws); wb.save(FILENAME); wb.close()
    except Exception as e:
        print("Error fetching ETF inflows:", e)
        stub = {"now": None, "yesterday": None, "week": None, "month": None}
        upsert_etf_inflows_row(FILENAME, SHEET_NAME, 16, stub)
        wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
        c_cell = ws.cell(row=16, column=3); c_cell.font = Font(size=13)
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_column_styles(ws); wb.save(FILENAME); wb.close()

    # CBBI overall (row 17) + components (rows 18+)
    try:
        today_str, cbbi_now, cbbi_metrics_today, cbbi_cache = update_cbbi_cache_for_today()
    except Exception as e:
        print("Error scraping CBBI page:", e)
        cbbi_cache = _load_cbbi_cache()

    cbbi_series = compute_cbbi_series_from_cache(cbbi_cache)
    upsert_cbbi_row(FILENAME, SHEET_NAME, 17, cbbi_series, weight=3)

    # Components start at row 18
    base_row = 18
    for idx, label in enumerate(CBBI_COMPONENT_LABELS):
        series = compute_cbbi_component_series_from_cache(cbbi_cache, label)
        row_number = base_row + idx
        upsert_cbbi_component_row(FILENAME, SHEET_NAME, row_number, label, series, weight=1)

    wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
    apply_column_styles(ws)
    wb.save(FILENAME); wb.close()

    # Finalize: always update A2 summary (weighted)
    wb = load_workbook(FILENAME); ws = wb[SHEET_NAME]
    update_risk_summary(ws)
    apply_column_styles(ws)
    wb.save(FILENAME); wb.close()

    print("Dashboard generation complete ✅")
